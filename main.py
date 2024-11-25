import os
import time
import pickle
from google_drive_downloader import GoogleDriveDownloader as gdd
from PIL import Image
import numpy as np
from svm import train_svm
from naive_bayes import train_naive_bayes
from decision_tree import train_decision_tree
from logistic_regression import train_logistic_regression
from neural_network import train_neural_network
from sklearn.model_selection import train_test_split
import pandas as pd
import sys
import itertools
import threading
import uuid
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

import os
import matplotlib.pyplot as plt
# Configuración de los algoritmos y sus respectivos entrenadores
algorithms = {
    'SVM': train_svm,
    'Naive Bayes': train_naive_bayes,
    'Decision Tree': train_decision_tree,
    'Logistic Regression': train_logistic_regression,
    'Neural Network': train_neural_network,
}

# Animación de carga mejorada con emojis
def loading_animation(message="Cargando"):
    emojis = ['🚀', '🌌', '🌕', '✨']
    for char in itertools.cycle(emojis):
        if loading_done:
            sys.stdout.write(f'\r{message} ✅ Completado!          \n')  # Asegura que se limpia la línea
            sys.stdout.flush()
            break
        sys.stdout.write(f'\r{message} {char}     ')  # Agrega espacios para limpiar la línea
        sys.stdout.flush()
        time.sleep(0.3)

# Función para cargar y preprocesar datos desde Google Drive o una carpeta local
def load_data(path, source='local'):
    global loading_done  # Variable para detener la animación
    data = []
    labels = []
    
    if source == 'drive':
        gdd.download_file_from_google_drive(file_id=path, dest_path='./temp_data', unzip=True)
        path = './temp_data'
    
    # Iniciar animación en un hilo separado
    loading_done = False
    thread = threading.Thread(target=loading_animation, args=("🔄 Procesando imágenes",))
    thread.start()
    
    try:
        for folder_name in os.listdir(path):
            folder_path = os.path.join(path, folder_name)
            if os.path.isdir(folder_path):
                for img_file in os.listdir(folder_path):
                    img_path = os.path.join(folder_path, img_file)
                    try:
                        img = Image.open(img_path).convert('RGB')
                        img = img.resize((224, 224))
                        data.append(np.array(img).flatten())
                        labels.append(folder_name)
                    except Exception as e:
                        print(f"\n👎  No se pudo cargar la imagen {img_file}: {e}")
    finally:
        loading_done = True
        thread.join()  # Esperar que la animación termine

    # Convertir listas a arrays numpy
    data = np.array(data)
    labels = np.array(labels)

    # Dividir datos en entrenamiento y prueba
    X_train, X_test, y_train, y_test = train_test_split(data, labels, test_size=0.3, random_state=42)

    return X_train, X_test, y_train, y_test

# Entrenamiento y evaluación de cada algoritmo con loading específico y guardado de modelos
def evaluate_algorithms(X_train, X_test, y_train, y_test, dataset_name):
    results = {}
    best_model = None
    best_accuracy = 0.0
    model_dir = os.path.join("entrenamiento", dataset_name)
    os.makedirs(model_dir, exist_ok=True)

    for name, train_func in algorithms.items():
        # Loading para cada algoritmo con emojis
        global loading_done
        loading_done = False
        print(f"🧠 Entrenando {name}...")
        thread = threading.Thread(target=loading_animation, args=(f"🚀 Entrenando {name}",))
        thread.start()
        
        try:
            # Entrenar y evaluar el modelo
            model_result = train_func(X_train, y_train, X_test, y_test)
            results[name] = model_result
            
            # Guardar el modelo entrenado en la carpeta correspondiente
            model_path = os.path.join(model_dir, f"{name}.pkl")
            with open(model_path, 'wb') as model_file:
                pickle.dump(model_result['model'], model_file)
            
            # Identificar el mejor modelo basado en la precisión
            if model_result['accuracy'] > best_accuracy:
                best_accuracy = model_result['accuracy']
                best_model = model_path
        finally:
            loading_done = True
            thread.join()  # Esperar que el loading termine
            print(f"✅ {name} completado.")
    
    return results, best_model

# Generación del reporte final en un archivo Excel
def generate_report(results, dataset_name):
    report_data = []
    for model_name, metrics in results.items():
        report_data.append({
            'Model': model_name,
            'Accuracy': metrics['accuracy'],
            'Precision': metrics['precision'],
            'Recall': metrics['recall'],
            'F1 Score': metrics['f1_score'],
            'AUC': metrics['auc'],
            'CPU Usage (%)': metrics['cpu_usage'],
            'Execution Time (s)': metrics['execution_time']
        })

    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.join("resultados", dataset_name)
    excel_dir = os.path.join(output_dir, "excel")
    images_dir = os.path.join(output_dir, "imagenes")
    
    # Crear las carpetas si no existen
    os.makedirs(excel_dir, exist_ok=True)
    os.makedirs(images_dir, exist_ok=True)

    # Guardar el reporte en Excel
    filename = f'Resultado_ml_{dataset_name}_{current_time}.xlsx'
    filepath = os.path.join(excel_dir, filename)

    df_report = pd.DataFrame(report_data)
    best_row_index = df_report['Accuracy'].idxmax()

    wb = Workbook()
    ws = wb.active
    ws.title = "Model Report"
    
    for r_idx, row in enumerate(dataframe_to_rows(df_report, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4F81BD")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center")
                if r_idx == best_row_index + 2:
                    cell.fill = PatternFill("solid", fgColor="90EE90")

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filepath)
    print(f"📊 Reporte generado: {filepath}")

    # Abrir el archivo automáticamente
    os.startfile(filepath)
    
    # Generar gráficos de barras para cada métrica
    metrics = ['Accuracy', 'Precision', 'Recall', 'F1 Score', 'AUC']
    for metric in metrics:
        plt.figure(figsize=(10, 6))
        plt.bar(df_report['Model'], df_report[metric], color='skyblue')
        plt.title(f'{metric} por Modelo')
        plt.xlabel('Modelo')
        plt.ylabel(metric)
        plt.ylim(0, 1)  # Suponiendo que las métricas están en el rango de 0 a 1
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        # Guardar el gráfico como archivo de imagen en la carpeta de imágenes
        chart_path = os.path.join(images_dir, f'{metric}_chart_{dataset_name}_{current_time}.png')
        plt.savefig(chart_path)
        print(f"📊 Gráfico {metric} guardado: {chart_path}")

        
        

# Función para clasificar una nueva imagen usando el mejor modelo
def classify_image(image_path, model_path):
    with open(model_path, 'rb') as file:
        model = pickle.load(file)
    
    img = Image.open(image_path).convert('RGB')
    img = img.resize((224, 224))
    img_array = np.array(img).flatten().reshape(1, -1)  # Preparar imagen para predicción
    prediction = model.predict(img_array)
    return prediction[0]

# Ejecución completa del sistema
def main(input_path, dataset_name, source='local'):
    X_train, X_test, y_train, y_test = load_data(input_path, source)
    results, best_model = evaluate_algorithms(X_train, X_test, y_train, y_test, dataset_name)
    generate_report(results, dataset_name)
    print(f"El mejor modelo entrenado se encuentra en: {best_model}")



# Entrenamiento
# main("C:\\Users\\rfrey\\Documents\\console_ml\\dataset", 'Dataset_de_Estrias', source='local')
# main("link drive abierto", 'Dataset_de_Estrias', source='drive')

# Clasificar una nueva imagen
# Ejemplo de uso de clasificación
# image_class = classify_image("C:\\Users\\rfrey\\Documents\\console_ml\\imagen\\e204236c65.JPG", "entrenamiento/Dataset_de_Estrias/SVM.pkl")
# print(f"🧠 La clase de la imagen es: {image_class}")
