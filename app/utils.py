#app/utils.py
from flask import Flask
import os
import time
import pickle
from google_drive_downloader import GoogleDriveDownloader as gdd
from PIL import Image
import numpy as np
from database import *
from algoritmos.svm import train_svm
from algoritmos.naive_bayes import train_naive_bayes
from algoritmos.decision_tree import train_decision_tree
from algoritmos.logistic_regression import train_logistic_regression
from algoritmos.neural_network import train_neural_network
from sklearn.model_selection import train_test_split
import pandas as pd
import sys
import itertools
import threading
import uuid
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from fpdf import FPDF
from app import create_app
import os
import matplotlib.pyplot as plt
# Configuración de los algoritmos y sus respectivos entrenadores
algorithms = {
    'SVM': train_svm,
    'Naive Bayes': train_naive_bayes,
    'Decision Tree': train_decision_tree,
    'Logistic Regression': train_logistic_regression,
    'Neural Network': train_neural_network,
}
setup_database()
# Animación de carga mejorada con emojis
def loading_animation(message="Cargando"):
    emojis = ['🧠', '⚙️', '🧠', '⚙️']
    for char in itertools.cycle(emojis):
        if loading_done:
            sys.stdout.write(f'\r{message} ✅ Completado!          \n')  # Asegura que se limpia la línea
            sys.stdout.flush()
            break
        sys.stdout.write(f'\r{message} {char}     ')  # Agrega espacios para limpiar la línea
        sys.stdout.flush()
        time.sleep(0.3)

# Función para cargar y preprocesar datos desde Google Drive o una carpeta local
def load_data(path, source='local'):
    global loading_done  # Variable para detener la animación
    data = []
    labels = []
    
    if source == 'drive':
        gdd.download_file_from_google_drive(file_id=path, dest_path='./temp_data', unzip=True)
        path = './temp_data'
    
    # Iniciar animación en un hilo separado
    loading_done = False
    thread = threading.Thread(target=loading_animation, args=("🧠 Procesando imágenes",))
    thread.start()
    
    try:
        for folder_name in os.listdir(path):
            folder_path = os.path.join(path, folder_name)
            if os.path.isdir(folder_path):
                for img_file in os.listdir(folder_path):
                    img_path = os.path.join(folder_path, img_file)
                    try:
                        img = Image.open(img_path).convert('RGB')
                        img = img.resize((224, 224))
                        data.append(np.array(img).flatten())
                        labels.append(folder_name)
                    except Exception as e:
                        print(f"\n👎  No se pudo cargar la imagen {img_file}: {e}")
    finally:
        loading_done = True
        thread.join()  # Esperar que la animación termine

    # Convertir listas a arrays numpy
    data = np.array(data)
    labels = np.array(labels)

    # Dividir datos en entrenamiento y prueba
    X_train, X_test, y_train, y_test = train_test_split(data, labels, test_size=0.3, random_state=42)

    return X_train, X_test, y_train, y_test

# Entrenamiento y evaluación de cada algoritmo con loading específico y guardado de modelos
def evaluate_algorithms(X_train, X_test, y_train, y_test, dataset_name, selected_algorithms):
    results = {}
    best_model = None
    best_accuracy = 0.0
    model_dir = os.path.join("entrenamiento", dataset_name)
    os.makedirs(model_dir, exist_ok=True)

    for name, train_func in algorithms.items():
        if name not in selected_algorithms:  # Saltar los no seleccionados
            continue
        
        global loading_done
        loading_done = False
        print(f"🧠 Entrenando {name}...")
        thread = threading.Thread(target=loading_animation, args=(f"⚙️ Entrenando {name}",))
        thread.start()
        
        try:
            model_result = train_func(X_train, y_train, X_test, y_test)
            results[name] = model_result
            
            # Guardar el modelo entrenado
            model_path = os.path.join(model_dir, f"{name}.pkl")
            with open(model_path, 'wb') as model_file:
                pickle.dump(model_result['model'], model_file)
            
            # Guardar información del modelo en la base de datos
            save_training_to_db(name, dataset_name, model_path)
            
            # Identificar el mejor modelo
            if model_result['accuracy'] > best_accuracy:
                best_accuracy = model_result['accuracy']
                best_model = model_path
        finally:
            loading_done = True
            thread.join()
            print(f"✅ {name} completado.")
    
    return results, best_model
# Generación del reporte final en un archivo Excel
def generate_report(results, dataset_name):
    if not os.path.exists("resultados"):
        os.makedirs("resultados")
    report_data = []
    for model_name, metrics in results.items():
        report_data.append({
            'Model': model_name,
            'Accuracy': metrics['accuracy'],
            'Precision': metrics['precision'],
            'Recall': metrics['recall'],
            'F1 Score': metrics['f1_score'],
            'AUC': metrics['auc'],
            'CPU Usage (%)': metrics['cpu_usage'],
            'Execution Time (s)': metrics['execution_time']
        })

    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = os.path.join("resultados", dataset_name)
    excel_dir = os.path.join(output_dir, "excel")
    images_dir = os.path.join(output_dir, "imagenes")
    
    os.makedirs(excel_dir, exist_ok=True)
    os.makedirs(images_dir, exist_ok=True)

    filename = f'Resultado_ml_{dataset_name}.xlsx'
    filepath = os.path.join(excel_dir, filename)

    df_report = pd.DataFrame(report_data)
    best_row_index = df_report['Accuracy'].idxmax()

    wb = Workbook()
    ws = wb.active
    ws.title = "Model Report"
    
    for r_idx, row in enumerate(dataframe_to_rows(df_report, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="4F81BD")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center")
                if r_idx == best_row_index + 2:
                    cell.fill = PatternFill("solid", fgColor="90EE90")

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(filepath)
    print(f"📊 Reporte generado: {filepath}")
    # Generar gráficos de barras para cada métrica
    metrics = ['Accuracy', 'Precision', 'Recall', 'F1 Score', 'AUC']
    # Abrir el archivo automáticamente
    for metric in metrics:
        plt.figure(figsize=(10, 6))
        plt.bar(df_report['Model'], df_report[metric], color='skyblue')
        plt.title(f'{metric} por Modelo')
        plt.xlabel('Modelo')
        plt.ylabel(metric)
        plt.ylim(0, 1)  # Suponiendo que las métricas están en el rango de 0 a 1
        plt.xticks(rotation=45)
        plt.tight_layout()
        
        # Guardar el gráfico como archivo de imagen en la carpeta de imágenes
        chart_path = os.path.join(images_dir, f'{metric}_chart_{dataset_name}.png')
        plt.savefig(chart_path)
        print(f"📊 Gráfico {metric} guardado: {chart_path}")
    # Generar reporte en PDF
    generate_pdf_report(images_dir, dataset_name)
    pdf_filename = f"Reporte_{dataset_name}.pdf"
    pdf_filepath = os.path.join(images_dir, pdf_filename)

    # Guardar información del dataset en la base de datos
    save_dataset_to_db(dataset_name, filepath, pdf_filepath)    

# Función para clasificar una nueva imagen usando el mejor modelo
def classify_image(image_path, model_path):
    print(f"[DEBUG] Ruta del modelo cargado: {model_path}")  # Depuración
    print(f"[DEBUG] Ruta de la imagen cargada: {image_path}")  # Depuración

    # Validar que la ruta del modelo existe
    if not os.path.exists(model_path):
        raise FileNotFoundError(f"El archivo del modelo no existe: {model_path}")

    # Validar que la imagen existe
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"El archivo de la imagen no existe: {image_path}")

    # Cargar el modelo desde la ruta
    with open(model_path, 'rb') as file:
        model = pickle.load(file)
        print(f"[DEBUG] Modelo cargado exitosamente desde: {model_path}")  # Depuración

    # Procesar la imagen
    img = Image.open(image_path).convert('RGB')
    img = img.resize((224, 224))
    img_array = np.array(img).flatten().reshape(1, -1)
    print(f"[DEBUG] Imagen procesada para predicción.")  # Depuración

    # Clasificar la imagen
    prediction = model.predict(img_array)
    print(f"[DEBUG] Predicción realizada: {prediction[0]}")  # Depuración
    return prediction[0]


def generate_pdf_report(images_dir, dataset_name):
    """
    Genera un PDF que combina gráficos e imágenes en un archivo único.
    """
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    pdf_filename = f"Reporte_{dataset_name}.pdf"
    pdf_filepath = os.path.join(images_dir, pdf_filename)
    
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    
    # Configuración general de la fuente
    pdf.set_font("Arial", size=12)
    
    # Agregar la portada
    pdf.add_page()
    pdf.set_font("Arial", size=16, style="B")
    pdf.cell(0, 10, f"Reporte de Resultados: {dataset_name}", ln=True, align='C')
    pdf.set_font("Arial", size=12)
    pdf.ln(10)
    pdf.cell(0, 10, f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", ln=True, align='C')
    pdf.ln(20)
    
    images = sorted([file for file in os.listdir(images_dir) if file.endswith(".png")])
    image_count = 0

    for file in images:
        img_path = os.path.join(images_dir, file)
        if image_count % 2 == 0:  # Si es el primer gráfico de la página
            if image_count > 0:  # Añadir una nueva página después de los dos gráficos anteriores
                pdf.add_page()
            y_offset = 30  # Posición inicial del primer gráfico
        else:  # Si es el segundo gráfico
            y_offset = 150  # Posición para el segundo gráfico

        # Título del gráfico
        pdf.set_font("Arial", size=14, style="B")
        pdf.image(img_path, x=10, y=y_offset, w=190)
        pdf.ln(10)

        image_count += 1

    # Guardar el PDF
    pdf.output(pdf_filepath)
    print(f"📄 PDF generado: {pdf_filepath}")
    
# Ejecución completa del sistema
def main(input_path, dataset_name, source='local', selected_algorithms=None):
    if selected_algorithms is None:
        selected_algorithms = list(algorithms.keys())
    
    X_train, X_test, y_train, y_test = load_data(input_path, source)
    results, best_model = evaluate_algorithms(X_train, X_test, y_train, y_test, dataset_name, selected_algorithms)

    # Generar reportes
    generate_report(results, dataset_name)
    print(f"El mejor modelo entrenado se encuentra en: {best_model}")